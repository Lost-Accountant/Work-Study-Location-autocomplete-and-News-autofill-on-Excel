import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#load main worksheet
excel_file = 'Kevin-coded-edited.xlsx'
workbook = load_workbook(filename= excel_file)
sheet = workbook.active

#dictionaries for agents
ref_file = 'LAND - Variable dictionary.xlsx'
ref_book = load_workbook(filename= ref_file)
ref_sheet = ref_book.active

row = 2
col = 2
col_ch = 'B'
agents = {}
#put keys into dict agents
while row <9:
    agents[ref_sheet['A' + str(row)].value] = []
    row +=1

row = 2

while row <9:
    while col < len(ref_sheet[row]):
        agents[ref_sheet['A'+str(row)].value].append(ref_sheet[col_ch+str(row)].value)
        col = col + 1
        col_ch = chr(ord(col_ch) + 1)    
    row = row + 1
    col = 2
    col_ch ='B'

print(agents)

#clean Non from agents
res_agents = {}
for keys in agents:
    res_agents[keys] = []
    for value in agents[keys]:
        if value is not None:
            res_agents[keys].append(value)

agents = res_agents
print(agents)



#c1 = sheet['E3'].value
#c2 = sheet['E'+str(3)].value

#print(c1==c2)

