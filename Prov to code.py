from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl

# load location code
ref_file = 'Province code.xlsx'
ref_book = load_workbook(filename=ref_file)
ref_sheet = ref_book.active

# Chinese name as key, code as value
code = {}
for row in range(1, ref_sheet.max_row + 1):
    code[ref_sheet['C'+ str(row)].value] = ref_sheet['A' + str(row)].value

# load main book to work on (here is copy)
main_file = 'WQW－Ruiyang Group－Kevin - Copy.xlsx'
main_book = load_workbook(filename=main_file)
main_sheet = main_book.active




for row in range(3, main_sheet.max_row + 1):
    # work on location I
    main_sheet['G'+str(row)] = code[main_sheet['I'+str(row)].value]
    # work on location II
    main_sheet['H'+str(row)] = code[main_sheet['J'+str(row)].value]

#print(code[main_sheet['I'+str(row)].value])
#print(main_sheet.cell(row = 7, column = row).value)
main_book.save(filename=main_file)