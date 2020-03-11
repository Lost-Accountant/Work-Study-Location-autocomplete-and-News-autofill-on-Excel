from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def load_html():
    file_name = input('what is the html name')
    html = r"C:\Users\gxsgt\Desktop\Violent Land\Hong Kong Protest\{}.html".format(file_name)

#html = r"C:\Users\gxsgt\Desktop\Violent Land\Hong Kong Protest\Factiva.html"

    with open(html, encoding='utf8') as f:
        html_string = f.read()

    soup = BeautifulSoup(html_string, 'html.parser')
    return soup


def extract_headline(each_article):
    headlines = each_article.find_all('span', {'class':"enHeadline"})
    headline = [headline.get_text() for headline in headlines]
    return headline[0]


def extract_author(each_article):
    authors = each_article.find_all('div', {'class':"author"})
    author = [author.get_text() for author in authors]
    if author == []:
        return 'NA'
    else:
        return author[0]


def extract_date(each_tag):
    if 'words' in each_tag[4].get_text():
        return each_tag[5].get_text()

    elif 'South' in each_tag[4].get_text():
        return each_tag[3].get_text()

    elif 'scmp.com' in each_tag[4].get_text():
        return each_tag[3].get_text()

    else:
        return each_tag[4].get_text()


def extract_id(each_tag):

    return each_tag[-1].get_text().replace('Document ','')


def extract_body(each_tag):
    paragraph = ''
    for sentence in each_tag:
        paragraph = paragraph + sentence.get_text() + "\n"

    return paragraph


def get_divtag(each_article):
    divtag = each_article.find_all('div', {'class': "article enArticle"})
    return divtag


def get_articles(soup):
    articles = soup.find_all('div', id =lambda x: x and x.startswith('article'))
    return articles

def filter(each_tag):
    filter_list = ['economy','lif','my take','spo','biz','opinion']
    interested = True
    for topic in filter_list:
        for tag in each_tag:
            if topic in tag.get_text().lower():
                interested = False

    return interested

def auto_fill():
    wb = Workbook()
    ws = wb.active

    # fill in column names
    col_names = ['ID','Headline','Author','Date','Text']
    for col_num in range(len(col_names)):
        ws.cell(row=1, column=col_num + 1).value = col_names[col_num]

    # create articles
    articles = get_articles(load_html())

    # set starting point
    n_row = 2

    # iterate through all articles
    for each_article in articles:

        # extract divtag for the rest
        divtag = get_divtag(each_article)

        # extract null_tag
        for tag in divtag:
            # not sure why loop needed here but it doesn't work if without
            each_p_tag = tag.find_all('p')
            # split divtag to each tag with no signature for the rest
            each_div_tag = tag.find_all('div')

            interested = filter(each_div_tag)

        # data entry only if interested topics
        if interested:
            # key in headline
            ws['B' + str(n_row)] = extract_headline(each_article)

            # key in author
            ws['C' + str(n_row)] = extract_author(each_article)

            # key in main body
            ws['E' + str(n_row)] =  extract_body(each_p_tag)

            # key in ID
            ws['A' + str(n_row)] = extract_id(each_p_tag)

            # key in date
            ws['D' + str(n_row)] = extract_date(each_div_tag)

            # increment
            n_row += 1

    wb.save('C:\\Users\\gxsgt\\Desktop\\Violent Land\\Hong Kong Protest\\SCMP.xlsx')
    return


if __name__ == "__main__":
    auto_fill()


# TO DO: ability to write on existing excel file
