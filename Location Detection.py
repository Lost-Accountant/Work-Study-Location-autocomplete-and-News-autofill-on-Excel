import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# load locations
file = open('pcas.json', encoding="utf8")

location = json.load(file)

# test how to read json
# print(location['北京市']['市辖区']['东城区'])

# load main worksheet
main_file = 'WQW－Ruiyang Group－Kevin.xlsx'
main_book = load_workbook(filename=main_file)
main_sheet = main_book.active

# find province in a cell

def search_prov(cellnum):
    found_p = []
    for province in location:
        # found province
        if province in main_sheet[cellnum].value:
            found_p.append(province)
        
    # in case no province listed
    if found_p == []:
        for province in location:
            found_p.append(province)

    return found_p

# find city from found province
def search_city(found_prov, cellnum):
    found_pc = {}
    for prov in found_prov:
        found_pc[prov] = {}
        for city in location[prov]:
            if city in main_sheet[cellnum].value:
                found_pc[prov][city] = {}
        
        # in case no city was found
        if found_pc[prov] == {}:
            for city in location[prov]:
                found_pc[prov][city] = {}
    return found_pc

# find district from city
def search_district(found_pc, cellnum):
    found_pcd = found_pc
    for prov in found_pc:
        for city in found_pc[prov]:
            for dist in location[prov][city]:
                if dist in main_sheet[cellnum].value:
                    found_pcd[prov][city][dist] = []

    # if at this point, still couldn't find one, probably time to give up
    # delete no hits on districts
    for prov in list(found_pcd.keys()):
       for city in list(found_pcd[prov].keys()):
            if found_pcd[prov][city] == {}:
                del found_pcd[prov][city]
    
    # delete no hits on provinces
    for prov in list(found_pcd.keys()):
        if found_pcd[prov] == {}:
            del found_pcd[prov]

    return found_pcd

def search_town(found_pcd, cellnum):
    found_pcdt = found_pcd
    for prov in found_pcd:
        for city in found_pcd[prov]:
            for dist in found_pcd[prov][city]:
                for town in location[prov][city][dist]:
                    if town in main_sheet[cellnum].value:
                        found_pcdt[prov][city][dist].append(town)
    
    return found_pcdt

cellnum = 'C57'

a1 = search_town(search_district(search_city(search_prov(cellnum), cellnum), cellnum), cellnum)

print(a1)

output = open('location output.txt','w', encoding='utf-8')

start = 3

for row in range(start, main_sheet.max_row + 1):
    output.write('Row #' + str(row) + '\n')
    found = search_town(search_district(search_city(search_prov('C' + str(row)), 'C' + str(row)), 'C' + str(row)), 'C' + str(row))
    for prov in found:
        for city in found[prov]:
            for dist in found[prov][city]:
                if found[prov][city][dist] != []:
                    for town in found[prov][city][dist]:
                        output.write(prov +','+city+','+dist+','+town+'\n')
                else:
                    output.write(prov +','+city+','+dist+'\n')
    output.write('\n')

output.close()
